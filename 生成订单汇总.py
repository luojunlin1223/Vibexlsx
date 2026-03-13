"""
从 Sheet A（订单明细）自动生成 Sheet B（按产品线×国家汇总）。
带完整 GUI 界面，支持自动更新。
"""

import sys
import os
import threading
import json
import subprocess
from collections import defaultdict
from urllib.request import urlopen, Request
from urllib.error import URLError
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

VERSION = "1.0.0"
GITHUB_REPO = "luojunlin1223/Vibexlsx"
UPDATE_URL = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
EXE_ASSET_NAME = "订单汇总生成器.exe"

# ========== 固定配置 ==========

COUNTRY_COL_MAP = {
    "India": 2,
    "Singapore": 3,
    "Vietnam": 4,
    "Philippines": 5,
    "Indonesia": 6,
    "Thailand": 7,
    "Malaysia": 8,
    "Japan": 9,
    "Korea": 10,
    "New Zealand": 11,
    "Australia": 12,
    "Other": 14,
}
CIS_COL = 13
TOTAL_COL = 15

CIS_COUNTRIES = {
    "Russia", "Kazakhstan", "Uzbekistan", "Turkmenistan", "Tajikistan",
    "Kyrgyzstan", "Belarus", "Ukraine", "Moldova", "Armenia",
    "Azerbaijan", "Georgia",
}

GROUP_HEADERS = {
    2: "Flexible Packaging",
    9: "Product Integrity&Material Test",
    17: "Food&Beverage",
}

PRODUCT_ROW_MAP = {
    "SI Permeation": 3,
    "Oxysense": 4,
    "TMI": 5,
    "Buchel": 6,
    "Technidyne": 7,
    "Rayran": 8,
    "TME": 10,
    "SI Process-Semiconductor": 11,
    "SI Process-others": 12,
    "SpecMetrix": 13,
    "UTS": 14,
    "TQC Sheen": 15,
    "C&W": 16,
    "CMC-KUHNKE": 18,
    "Quality By Vision": 19,
    "Eagle Vision": 20,
    "XTS": 21,
    "Steinfurth": 22,
    "Torus": 23,
    "SI Headspace": 24,
    "3rd party": 25,
    "Service": 26,
}

TOTAL_ROW = 27
DATA_ROW_START = 3
DATA_ROW_END = 26

NUM_FMT = '_ * #,##0.00_ ;_ * \\-#,##0.00_ ;_ * "-"??_ ;_ @_ '
NUM_FMT_TOTAL = "0.000000"

COL_HEADERS = [
    "India", "Singapore", "Vietnam", "Philippines", "Indonesia",
    "Thailand", "Malaysia", "Japan", "Korea", "New Zealand",
    "Australia", "CIS Countries", "Other", "APAC(Total)",
]

# 列号 → 国家名（用于日志显示）
COL_TO_COUNTRY = {v: k for k, v in COUNTRY_COL_MAP.items()}
COL_TO_COUNTRY[CIS_COL] = "CIS Countries"


def get_country_col(country_name):
    if not country_name:
        return None
    country_name = country_name.strip()
    if country_name in COUNTRY_COL_MAP:
        return COUNTRY_COL_MAP[country_name]
    if country_name in CIS_COUNTRIES:
        return CIS_COL
    return COUNTRY_COL_MAP["Other"]


def read_sheet_a(ws):
    agg = defaultdict(float)
    warnings = []
    row_count = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        product = row[14].value   # O列
        country = row[11].value   # L列
        q5 = row[16].value        # Q列
        value = row[22].value     # W列

        if not product or value is None:
            continue

        # 跳过公式字符串等非数值
        if not isinstance(value, (int, float)):
            warnings.append(f"第{row[0].row}行 W列不是数值，已跳过")
            continue

        row_count += 1

        if q5 and str(q5).strip() == "Parts":
            product = "Service"
        else:
            product = str(product).strip()

        col = get_country_col(country)
        if col is None:
            continue

        if product not in PRODUCT_ROW_MAP:
            warnings.append(f"未知产品线 '{product}'，已跳过")
            continue

        agg[(product, col)] += value

    return agg, warnings, row_count


def write_sheet_b(wb, agg):
    if "Sheet B" in wb.sheetnames:
        del wb["Sheet B"]

    ws = wb.create_sheet("Sheet B")

    font_bold = Font(bold=True)
    font_normal = Font(bold=False)
    align_center = Alignment(horizontal="center")
    align_right = Alignment(horizontal="right")

    for i, header in enumerate(COL_HEADERS):
        cell = ws.cell(row=1, column=i + 2, value=header)
        cell.alignment = align_center
        if header == "APAC(Total)":
            cell.font = font_bold

    for row_num, title in GROUP_HEADERS.items():
        cell = ws.cell(row=row_num, column=1, value=title)
        cell.font = font_bold

    for product, row_num in PRODUCT_ROW_MAP.items():
        cell = ws.cell(row=row_num, column=1, value=product)
        if product in ("3rd party", "Service"):
            cell.font = font_bold
        else:
            cell.font = font_normal

    for (product, col), value in agg.items():
        row_num = PRODUCT_ROW_MAP[product]
        cell = ws.cell(row=row_num, column=col, value=value / 1000)
        cell.number_format = NUM_FMT
        cell.alignment = align_right

    b_letter = get_column_letter(2)
    n_letter = get_column_letter(14)
    for row_num in range(DATA_ROW_START, DATA_ROW_END + 1):
        if row_num in GROUP_HEADERS:
            continue
        formula = f"=SUM({b_letter}{row_num}:{n_letter}{row_num})"
        cell = ws.cell(row=row_num, column=TOTAL_COL, value=formula)
        if row_num == 25:
            cell.font = font_bold
            cell.number_format = NUM_FMT_TOTAL
        else:
            cell.number_format = NUM_FMT
        cell.alignment = align_right

    total_cell_a = ws.cell(row=TOTAL_ROW, column=1, value="Total")
    total_cell_a.font = font_bold
    for col in range(2, TOTAL_COL + 1):
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}{DATA_ROW_START}:{col_letter}{DATA_ROW_END})"
        cell = ws.cell(row=TOTAL_ROW, column=col, value=formula)
        cell.font = font_bold
        cell.number_format = NUM_FMT_TOTAL
        cell.alignment = align_right

    ws.column_dimensions["A"].width = 28
    for col in range(2, TOTAL_COL + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions[get_column_letter(TOTAL_COL)].width = 16


# ========== 自动更新 ==========

def _parse_version(v):
    """将版本字符串 'v1.2.3' 或 '1.2.3' 转为可比较的元组。"""
    v = v.lstrip("vV")
    parts = []
    for p in v.split("."):
        try:
            parts.append(int(p))
        except ValueError:
            parts.append(0)
    return tuple(parts)


def check_update():
    """
    检查 GitHub 最新 release，返回 (has_update, latest_version, download_url, release_notes)。
    如果检查失败返回 (False, None, None, None)。
    """
    try:
        req = Request(UPDATE_URL, headers={"Accept": "application/vnd.github.v3+json"})
        with urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        latest_tag = data.get("tag_name", "")
        release_notes = data.get("body", "") or ""

        if not latest_tag:
            return False, None, None, None

        if _parse_version(latest_tag) <= _parse_version(VERSION):
            return False, latest_tag, None, release_notes

        # 查找 exe 资产下载链接
        download_url = None
        for asset in data.get("assets", []):
            if asset.get("name") == EXE_ASSET_NAME:
                download_url = asset.get("browser_download_url")
                break

        return True, latest_tag, download_url, release_notes
    except Exception:
        return False, None, None, None


def download_and_replace(download_url, progress_callback=None):
    """
    下载新版 exe 并替换当前文件。
    progress_callback(downloaded_mb, total_mb) 用于报告进度。
    返回 (success, message)。
    """
    if getattr(sys, 'frozen', False):
        current_exe = sys.executable
    else:
        return False, "仅打包后的 exe 支持自动更新。"

    old_exe = current_exe + ".old"

    try:
        # 下载新版本到临时文件
        tmp_exe = current_exe + ".tmp"
        req = Request(download_url)
        with urlopen(req, timeout=60) as resp:
            total = int(resp.headers.get("Content-Length", 0))
            total_mb = total / (1024 * 1024) if total else 0
            downloaded = 0
            with open(tmp_exe, "wb") as f:
                while True:
                    chunk = resp.read(256 * 1024)
                    if not chunk:
                        break
                    f.write(chunk)
                    downloaded += len(chunk)
                    if progress_callback:
                        progress_callback(downloaded / (1024 * 1024), total_mb)

        # Windows 允许重命名正在运行的 exe
        if os.path.exists(old_exe):
            os.remove(old_exe)
        os.rename(current_exe, old_exe)
        os.rename(tmp_exe, current_exe)

        return True, current_exe

    except Exception as e:
        # 回滚
        if os.path.exists(old_exe) and not os.path.exists(current_exe):
            os.rename(old_exe, current_exe)
        tmp_exe = current_exe + ".tmp"
        if os.path.exists(tmp_exe):
            os.remove(tmp_exe)
        return False, f"更新失败：{e}"


# ========== GUI ==========

class App:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title(f"订单汇总生成器 v{VERSION}")
        self.root.geometry("700x560")
        self.root.resizable(False, False)

        self.input_path = tk.StringVar()
        self.output_path = tk.StringVar()

        self._build_ui()
        # 启动后台检查更新
        threading.Thread(target=self._check_update_background, daemon=True).start()

    def _build_ui(self):
        root = self.root

        # ---------- 顶部栏：标题 + 更新按钮 ----------
        top_frame = tk.Frame(root)
        top_frame.pack(fill="x", padx=20, pady=(12, 0))

        title_label = tk.Label(top_frame, text="本月已收到订单汇总", font=("Microsoft YaHei", 16, "bold"))
        title_label.pack(side="left", padx=(0, 10))

        self.btn_update = tk.Button(top_frame, text="检查更新", font=("Microsoft YaHei", 9),
                                     command=self._on_check_update, width=10)
        self.btn_update.pack(side="right")

        self.version_label = tk.Label(top_frame, text=f"v{VERSION}", font=("Microsoft YaHei", 9), fg="#999999")
        self.version_label.pack(side="right", padx=(0, 8))

        subtitle = tk.Label(root, text="从 Sheet A 自动生成 Sheet B 汇总表", font=("Microsoft YaHei", 10), fg="#666666")
        subtitle.pack(pady=(2, 10))

        # ---------- 输入文件 ----------
        frame_in = tk.LabelFrame(root, text="输入文件", font=("Microsoft YaHei", 10), padx=12, pady=8)
        frame_in.pack(fill="x", padx=20, pady=(0, 8))

        entry_in = tk.Entry(frame_in, textvariable=self.input_path, font=("Microsoft YaHei", 9), state="readonly")
        entry_in.pack(side="left", fill="x", expand=True, ipady=3)

        btn_browse_in = tk.Button(frame_in, text="浏览...", font=("Microsoft YaHei", 9), command=self._browse_input, width=8)
        btn_browse_in.pack(side="right", padx=(8, 0))

        # ---------- 输出文件 ----------
        frame_out = tk.LabelFrame(root, text="输出文件", font=("Microsoft YaHei", 10), padx=12, pady=8)
        frame_out.pack(fill="x", padx=20, pady=(0, 8))

        entry_out = tk.Entry(frame_out, textvariable=self.output_path, font=("Microsoft YaHei", 9), state="readonly")
        entry_out.pack(side="left", fill="x", expand=True, ipady=3)

        btn_browse_out = tk.Button(frame_out, text="浏览...", font=("Microsoft YaHei", 9), command=self._browse_output, width=8)
        btn_browse_out.pack(side="right", padx=(8, 0))

        # ---------- 生成按钮 ----------
        self.btn_run = tk.Button(root, text="生 成", font=("Microsoft YaHei", 12, "bold"),
                                 bg="#4CAF50", fg="white", activebackground="#45a049",
                                 width=14, height=1, command=self._run)
        self.btn_run.pack(pady=10)

        # ---------- 日志区域 ----------
        frame_log = tk.LabelFrame(root, text="处理日志", font=("Microsoft YaHei", 10), padx=12, pady=8)
        frame_log.pack(fill="both", expand=True, padx=20, pady=(0, 16))

        self.log_text = tk.Text(frame_log, font=("Consolas", 9), height=10, state="disabled", wrap="word")
        scrollbar = tk.Scrollbar(frame_log, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.log_text.pack(fill="both", expand=True)

    def _log(self, msg, tag=None):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", msg + "\n", tag)
        self.log_text.see("end")
        self.log_text.configure(state="disabled")
        self.root.update_idletasks()

    def _log_clear(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

    def _browse_input(self):
        path = filedialog.askopenfilename(
            title="选择订单汇总 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
        )
        if path:
            self.input_path.set(path)
            base, ext = os.path.splitext(path)
            self.output_path.set(f"{base}_output{ext}")

    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            title="保存输出文件",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if path:
            self.output_path.set(path)

    def _run(self):
        input_file = self.input_path.get()
        output_file = self.output_path.get()

        if not input_file:
            messagebox.showwarning("提示", "请先选择输入文件！")
            return
        if not output_file:
            messagebox.showwarning("提示", "请先设置输出文件路径！")
            return

        self._log_clear()
        self.btn_run.configure(state="disabled", text="处理中...")

        thread = threading.Thread(target=self._run_worker, args=(input_file, output_file), daemon=True)
        thread.start()

    def _schedule(self, func, *args):
        self.root.after(0, func, *args)

    def _reset_button(self):
        self.btn_run.configure(state="normal", text="生 成")

    def _run_worker(self, input_file, output_file):
        try:
            self._schedule(self._log, f"读取文件: {os.path.basename(input_file)}")

            wb = openpyxl.load_workbook(input_file, data_only=True)

            if "Sheet A" not in wb.sheetnames:
                self._schedule(self._log, "[错误] 找不到 'Sheet A' 工作表！")
                self._schedule(messagebox.showerror, "错误", "找不到 'Sheet A' 工作表！\n请确保输入文件包含名为 'Sheet A' 的工作表。")
                return

            ws_a = wb["Sheet A"]
            self._schedule(self._log, "正在解析 Sheet A 数据...")

            agg, warnings, row_count = read_sheet_a(ws_a)

            self._schedule(self._log, f"共读取 {row_count} 条有效数据行")
            self._schedule(self._log, f"汇总为 {len(agg)} 个产品×国家组合:")
            self._schedule(self._log, "")

            for (product, col), value in sorted(agg.items()):
                country_name = COL_TO_COUNTRY.get(col, f"列{col}")
                self._schedule(self._log, f"  {product:25s} | {country_name:15s} | {value:,.2f}")

            self._schedule(self._log, "")

            for w in set(warnings):
                self._schedule(self._log, f"[警告] {w}")

            self._schedule(self._log, "正在生成 Sheet B...")
            write_sheet_b(wb, agg)

            self._schedule(self._log, f"正在保存: {os.path.basename(output_file)}")
            wb.save(output_file)

            self._schedule(self._log, "")
            self._schedule(self._log, f"完成！输出文件: {output_file}")
            self._schedule(messagebox.showinfo, "完成", f"处理完成！\n\n输出文件：\n{output_file}")

        except Exception as e:
            self._schedule(self._log, f"\n[错误] {e}")
            self._schedule(messagebox.showerror, "错误", f"处理过程中出错：\n{e}")

        finally:
            self._schedule(self._reset_button)

    # ---------- 更新相关 ----------

    def _check_update_background(self):
        """启动时后台静默检查更新。"""
        has_update, latest, download_url, notes = check_update()
        if has_update and download_url:
            self._schedule(self._prompt_update, latest, download_url, notes)

    def _on_check_update(self):
        """用户手动点击检查更新。"""
        self.btn_update.configure(state="disabled", text="检查中...")
        threading.Thread(target=self._check_update_manual, daemon=True).start()

    def _check_update_manual(self):
        has_update, latest, download_url, notes = check_update()
        if has_update and download_url:
            self._schedule(self._prompt_update, latest, download_url, notes)
        elif has_update and not download_url:
            self._schedule(messagebox.showinfo, "更新", f"发现新版本 {latest}，但未找到可下载的 exe 文件。")
        else:
            self._schedule(messagebox.showinfo, "更新", f"当前已是最新版本 v{VERSION}")
        self._schedule(self._reset_update_button)

    def _reset_update_button(self):
        self.btn_update.configure(state="normal", text="检查更新")

    def _prompt_update(self, latest, download_url, notes):
        """弹窗询问用户是否更新。"""
        msg = f"发现新版本 {latest}（当前 v{VERSION}）\n"
        if notes:
            # 只显示前200字符的更新说明
            short_notes = notes[:200] + ("..." if len(notes) > 200 else "")
            msg += f"\n更新说明：\n{short_notes}\n"
        msg += "\n是否立即更新？"

        if messagebox.askyesno("发现新版本", msg):
            self.btn_update.configure(state="disabled", text="更新中...")
            self._log_clear()
            self._log(f"正在下载新版本 {latest} ...")
            threading.Thread(target=self._do_update, args=(download_url,), daemon=True).start()

    def _do_update(self, download_url):
        """在子线程中执行下载和替换。"""
        def on_progress(downloaded_mb, total_mb):
            if total_mb > 0:
                self._schedule(self._log_update_progress, downloaded_mb, total_mb)

        success, result = download_and_replace(download_url, progress_callback=on_progress)

        if success:
            self._schedule(self._log, "")
            self._schedule(self._log, "更新完成！即将重启应用...")
            self._schedule(self._restart_app, result)
        else:
            self._schedule(self._log, f"\n{result}")
            self._schedule(messagebox.showerror, "更新失败", result)
            self._schedule(self._reset_update_button)

    def _log_update_progress(self, downloaded_mb, total_mb):
        """更新日志区的下载进度（覆盖最后一行）。"""
        self.log_text.configure(state="normal")
        # 删除最后一行进度（如果有的话）
        last_line_start = self.log_text.index("end-2l linestart")
        last_line_text = self.log_text.get(last_line_start, "end-1c")
        if last_line_text.startswith("  下载进度"):
            self.log_text.delete(last_line_start, "end-1c")
        progress_text = f"  下载进度: {downloaded_mb:.1f} / {total_mb:.1f} MB ({downloaded_mb/total_mb*100:.0f}%)"
        self.log_text.insert("end", progress_text + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _restart_app(self, exe_path):
        """重启应用。"""
        # 启动新进程
        subprocess.Popen([exe_path])
        self.root.destroy()

    def run(self):
        self.root.mainloop()


def main():
    if len(sys.argv) >= 2:
        input_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) >= 3 else None
        if output_file is None:
            base, ext = os.path.splitext(input_file)
            output_file = f"{base}_output{ext}"
        try:
            wb = openpyxl.load_workbook(input_file, data_only=True)
            if "Sheet A" not in wb.sheetnames:
                print("错误：找不到 'Sheet A' 工作表")
                sys.exit(1)
            agg, warnings, row_count = read_sheet_a(wb["Sheet A"])
            print(f"读取 {row_count} 条数据，汇总为 {len(agg)} 个组合")
            for w in set(warnings):
                print(f"[警告] {w}")
            write_sheet_b(wb, agg)
            wb.save(output_file)
            print(f"输出文件: {output_file}")
        except Exception as e:
            print(f"错误：{e}")
            sys.exit(1)
    else:
        App().run()


if __name__ == "__main__":
    main()
